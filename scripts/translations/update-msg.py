import csv
import os
import re
import sys
import traceback
from openpyxl import load_workbook


def read_excel(file_path):
    """Read the excel file.

    :param file_path: Path of file excel.
    :return: List of row of update data.
    """
    wb = load_workbook(filename=file_path)
    ws = wb['resouce一覧']

    # get all rows that have data in column D.
    rows_update = []
    for wc in ws['D3:D{}'.format(ws.max_row)]:
        if wc[0].value and not re.search(r'^【.*】$', wc[0].value):
            rows_update.append({
                'row': wc[0].row,
                'data': ws['A{}:D{}'.format(wc[0].row, wc[0].row)]
            })
    return rows_update


def prepare_list_replace_msg(data, weko_path):
    """Prepare list of replace message with data from excel.

    :param data: Data need update from excel.
    :param weko_path: weko path.
    :return: .
    """
    list_line = []
    for row in data:
        _row = row['data'][0]

        module_name = _row[0].value.split('/')[1]
        path = '/{}/{}/translations'.format(module_name,
                                            module_name.replace('-', '_'))
        en_path = path + '/en/LC_MESSAGES/messages.po'
        jp_path = path + '/ja/LC_MESSAGES/messages.po'
        if _row[1].value.startswith('msgid'):
            list_line.append({
                'excel_row': row['row'],
                'po_path': en_path,
                'type': 'en',
                'msg_old': '"{}"'.format(_row[2].value) if _row[2].value else None,
                'msg_new': '"{}"'.format(_row[3].value),
                'module_name': module_name
            })
        elif _row[1].value.startswith('msgstr'):
            list_line.append({
                'excel_row': row['row'],
                'po_path': jp_path,
                'type': 'jp',
                'msg_old': '"{}"'.format(_row[2].value) if _row[2].value else None,
                'msg_new': '"{}"'.format(_row[3].value),
                'module_name': module_name
            })

    return list_line


def replace_messages_po(imput_data, weko_path):
    """Replace all message need to replace in po file.

    :param imput_data: Input data.
    :param weko_path: weko path.
    :return: List result.
    """
    result = {
        'done': [],
        'error': []
    }
    for idx, data in enumerate(imput_data, start=1):
        is_done = False
        file_content = None
        # read input file
        msg_file_path = weko_path + '/modules' + data['po_path']
        if not os.path.exists(msg_file_path):
            result['error'].append({
                'index': idx,
                'excel_row': data['excel_row'],
                'error': '{} is not existed.'.format(msg_file_path),
                'data': data
            })
            continue

        with open(msg_file_path, "rt") as f:
            file_content = f.read()
            if data['msg_old'] and data['msg_old'] in file_content:
                if data['msg_old'] == data['msg_new']:
                    result['error'].append({
                        'index': idx,
                        'excel_row': data['excel_row'],
                        'error': 'Data in C and D are similar.',
                        'data': data
                    })
                else:
                    if data.get('type') == 'en':
                        file_content = re.sub(
                            '\\nmsgid\s+{}\s*(\\n)+msgstr\s+"\w*"'.format(
                                re.escape(data['msg_old'])),
                            '\nmsgid {}\nmsgstr {}'.format(
                                data['msg_old'], data['msg_new']),
                            file_content
                        )
                    else:
                        file_content = file_content.replace(
                            data['msg_old'], data['msg_new'])
                    is_done = True
                    result['done'].append(data)
            else:
                result['error'].append({
                    'index': idx,
                    'excel_row': data['excel_row'],
                    'error': 'C is empty, but D have data.' if not data['msg_old'] else 'Data in C is not found.',
                    'data': data
                })

        if is_done:
            with open(msg_file_path, "wt") as f:
                f.write(file_content)
    return result


def build_mo_file(module_names):
    for module_name in module_names:
        os.system('cd /code/modules/{}'.format(module_name) +
                  ';python setup.py compile_catalog')


def print_result(result, weko_path):
    print(''.ljust(8), end='')
    print('Success: {} messages.'.format(len(result['done'])))
    print(''.ljust(16), end='')
    print('List rows in Excel are succeeded: ', end='')
    print([d['excel_row'] for d in result['done']])
    print('')
    print(''.ljust(8), end='')
    print('Fail: {} messages.'.format(len(result['error'])))
    # print error details.
    print('')
    print('Error Details:')
    print('NO'.ljust(5), end='')
    print('EXCEL ROW'.ljust(20), end='')
    print('ERROR'.ljust(120), end='')
    print('MESSAGE (C)')
    output_err_data_to_file = []
    for idx, err in enumerate(result['error'], start=1):
        print(str(idx).ljust(5), end='')
        print(str(err['excel_row']).ljust(20), end='')
        print(err['error'].ljust(120), end='')
        print(err['data']['msg_old'])
        output_err_data_to_file.append([
            idx, err['excel_row'],
            err['error'], err['data']['msg_old'], err['data']
        ])
    # print error to file
    with open(weko_path + '/replace_msg_error.csv', mode='w') as file:
        writer = csv.writer(file, delimiter=',',
                            quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['NO', 'EXCEL ROW',
                         'ERROR', 'MESSAGE (C)', 'DATA'])
        writer.writerows(output_err_data_to_file)
    print('')
    print('All error informations in {}/replace_msg_error.csv'.format(weko_path))


if __name__ == "__main__":
    print('----- START -----')
    try:
        file_path = sys.argv[1] if len(
            sys.argv) > 1 else 'scripts/update_msg/19309_チェック結果_202000428v2_整理前_yutaka_mhaya改.xlsx'
        weko_path = sys.argv[2] if len(
            sys.argv) > 2 else ''
        weko_path += '' if weko_path and weko_path[-1] == '/' else ''
        print('Step 1: Read the excel file: ' + file_path)
        if not os.path.exists(file_path):
            print('File {} is not existed!')
        else:
            excel_data = read_excel(file_path)
            print('')
            print('Step 2: Prepare list of replace messages.')
            replace_msgs = prepare_list_replace_msg(excel_data, weko_path)
            print('')
            print('Step 3: Replacing...')
            result = replace_messages_po(replace_msgs, weko_path)
            print('')
            print('Step 4: Building mo file...')
            build_mo_file(set([d['module_name'] for d in result['done']]))
            print('')
            print('Step 5: Results: {} messages.'.format(len(replace_msgs)))
            print_result(result, weko_path)
    except Exception as ex:
        traceback.print_exc()
    print('----- END -----')
